import pickle
import openpyxl

with open('result.pickle', 'rb') as f:
    data = pickle.load(f)

file = openpyxl.Workbook()

def get_one_sheet(file, dict_name, data):
    ws = file.create_sheet(dict_name)
    for i, element in enumerate(data):
        district, dicts = element['district'], element['data']
        for d in dicts:
            if d['dict'] == dict_name:
                attrs_count = len(d['attrs'])
                cell = ws.cell(row=1, column=i*(attrs_count+1)+1, value=district)

                for attr_index, attr in enumerate(d['attrs']):
                    cell = ws.cell(row=2, column=i*(attrs_count+1)+attr_index+1)
                    cell.value = attr
                for j, row in enumerate(d['vals']):
                    for val_index, val in enumerate(row):
                        cell = ws.cell(row=3+j, column=i*(attrs_count+1)+val_index+1)
                        cell.value = val
    return ws


dicts = []
with open('dict.txt', 'r',  encoding='utf-8') as f:
    lines = f.readlines()
    for line in lines:
        args = line.split('|')
        dicts.append({
            'table': args[0].strip(),
            'attr': args[1].strip(),
        })

for dict_item in dicts:
    ws = get_one_sheet(file, dict_item['table'], data)

file.save('result.xlsx')
