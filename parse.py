import cx_Oracle
import pickle
cx_Oracle.init_oracle_client(lib_dir=r"instantclient_21_9")

dicts = []
with open('dict.txt', 'r',  encoding='utf-8') as f:
    lines = f.readlines()
    for line in lines:
        args = line.split('|')
        dicts.append({
            'table': args[0].strip(),
            'attr': args[1].strip(),
        })

connections = []
with open('database.txt', 'r', encoding='utf-8') as f:
    lines = f.readlines()
    for line in lines:
        args = line.split('|')
        connections.append({
            'district': args[0].strip(),
            'host': args[1].strip(),
            'database': args[2].strip(),
            'username': args[3].strip(),
            'password': args[4].strip()
        })
print(dicts)
print(connections)
gdata = []
where = input('Введите условие с WHERE. Пример: WHERE ATTR = 57. \nЕсли условия нет, введите пробел.\n')

def parse(connection):
    data = []
    conn = cx_Oracle.connect(user=connection['username'], password=connection['password'],
                                    dsn=f"{connection['host']}/{connection['database']}")

    for d in dicts:
        with conn.cursor() as cursor:
            print(f"    start: {d['table']}")
            sort_field = d['attr'].split(',')[0]
            cursor.execute(f"""
                    SELECT {d['attr']}
                    FROM CBSC.{d['table']} 
                    {where}""")
            rows = cursor.fetchall()
            val_list = []
            # print(rows)
            for row in rows:
                val_list.append(row)
            data.append({
                'dict': d['table'],
                'attrs': d['attr'].split(','),
                'vals': val_list
            })
    return {
        'district': connection['district'],
        'data': data
    }

for connection in connections:
    data = None
    print(f"START PARSING: {connection['district']}")
    try:
        data = parse(connection)
        print('END')
        print('')
    except Exception as e:
        print(e)
    if data is None:
        continue
    gdata.append(data)


with open('result.pickle', 'wb') as f:
    pickle.dump(gdata, f)



import pickle
import openpyxl

with open('result.pickle', 'rb') as f:
    data = pickle.load(f)

file = openpyxl.Workbook()

def get_one_sheet(file, dict_name, data):
    ws = file.create_sheet(dict_name)
    for i, element in enumerate(data):
        district, dicts = element['district'], element['data']
        for d in dicts:
            if d['dict'] == dict_name:
                attrs_count = len(d['attrs'])
                cell = ws.cell(row=1, column=i*(attrs_count+1)+1, value=district)

                for attr_index, attr in enumerate(d['attrs']):
                    cell = ws.cell(row=2, column=i*(attrs_count+1)+attr_index+1)
                    cell.value = attr
                for j, row in enumerate(d['vals']):
                    for val_index, val in enumerate(row):
                        cell = ws.cell(row=3+j, column=i*(attrs_count+1)+val_index+1)
                        cell.value = val
    return ws


dicts = []
with open('dict.txt', 'r',  encoding='utf-8') as f:
    lines = f.readlines()
    for line in lines:
        args = line.split('|')
        dicts.append({
            'table': args[0].strip(),
            'attr': args[1].strip(),
        })

for dict_item in dicts:
    ws = get_one_sheet(file, dict_item['table'], data)

file.save('result.xlsx')

